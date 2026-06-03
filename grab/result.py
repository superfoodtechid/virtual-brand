from __future__ import annotations

import argparse
import os
from pathlib import Path

import pandas as pd
import requests
from dotenv import load_dotenv

# Load environment variables
load_dotenv()


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = BASE_DIR / "downloads" / "grab_transactions_3months_(01-02-26_to_30-04-26).csv"
DEFAULT_OUTPUT = BASE_DIR / "monthly_summary_wide.xlsx"
LAPORAN_DIR = BASE_DIR / "laporan"


def resolve_input_path(raw_path: str | None) -> Path:
	if raw_path:
		return Path(raw_path).expanduser().resolve()

	# Cari file terbaru di folder downloads
	downloads_dir = BASE_DIR / "downloads"
	if downloads_dir.exists():
		files = list(downloads_dir.glob("grab_transactions_*.csv"))
		if files:
			# Ambil file yang paling baru dimodifikasi
			latest_file = max(files, key=lambda p: p.stat().st_mtime)
			print(f"Menggunakan file terbaru: {latest_file.name}")
			return latest_file

	return DEFAULT_INPUT


def load_transactions(csv_path: Path) -> pd.DataFrame:
	if not csv_path.exists():
		raise FileNotFoundError(f"CSV file not found: {csv_path}")

	df = pd.read_csv(csv_path)

	required_columns = {"Created On", "Long Order ID", "Net Sales", "Category"}
	missing_columns = required_columns.difference(df.columns)
	if missing_columns:
		raise ValueError(f"Missing required columns: {', '.join(sorted(missing_columns))}")

	return df


def summarize_monthly(
	df: pd.DataFrame,
	username: str = None,
	date_start: pd.Timestamp = None,
	date_end: pd.Timestamp = None,
) -> pd.DataFrame:
	working = df.copy()

	# Parse date columns — use 'Created On' as primary date for filtering
	# 'Updated On' is the settlement/disbursement date and can fall outside the order range
	date_fmt = "%d %b %Y %I:%M %p"
	working["Created On"] = pd.to_datetime(working.get("Created On"), errors="coerce", format=date_fmt)
	working["Updated On"] = pd.to_datetime(working.get("Updated On"), errors="coerce", format=date_fmt)
	
	# Use Created On for filtering; fall back to Updated On if Created On is missing
	working["_filter_date"] = working["Created On"].fillna(working["Updated On"])

	working["Long Order ID"] = working["Long Order ID"].fillna("").astype(str).str.strip()
	working["Category"] = working["Category"].fillna("").astype(str).str.strip().str.casefold()
	working["Net Sales"] = pd.to_numeric(working["Net Sales"], errors="coerce").fillna(0)
	working["Status"] = working["Status"].fillna("").astype(str).str.strip().str.casefold()

	valid_long_order_id = working["Long Order ID"].str.match(r"^[A-Za-z0-9-]+$", na=False)
	
	# An order is valid if it has a proper ID and is a Payment OR an Adjustment
	# We exclude Cancelled orders
	is_order_category = working["Category"].isin(["payment", "adjustment"])
	is_not_cancelled = working["Status"].ne("cancelled")
	
	valid_orders = working.loc[valid_long_order_id & is_order_category & is_not_cancelled].copy()
	valid_orders = valid_orders.loc[valid_orders["_filter_date"].notna()].copy()
	if date_start is not None:
		valid_orders = valid_orders.loc[valid_orders["_filter_date"] >= date_start].copy()
	if date_end is not None:
		valid_orders = valid_orders.loc[valid_orders["_filter_date"] <= date_end].copy()
	valid_orders["Month"] = valid_orders["_filter_date"].dt.to_period("M").dt.to_timestamp()

	summary = (
		valid_orders.groupby("Month", as_index=False)
		.agg(
			Order_Count=("Long Order ID", "count"), # Count rows to match stakeholder requirement
			Omzet_Net_Sales=("Net Sales", "sum"),
		)
		.sort_values("Month")
		.reset_index(drop=True)
	)

	summary.insert(0, "Username", username or os.getenv("GRAB_USERNAME", "unknown"))

	return summary


def format_rupiah(value: float | int | None) -> str:
	if pd.isna(value):
		return "Rp0"

	number = round(float(value))
	return f"Rp{number:,.0f}".replace(",", ".")


def summarize_wide(
	df: pd.DataFrame,
	username: str = None,
	date_start: pd.Timestamp = None,
	date_end: pd.Timestamp = None,
) -> pd.DataFrame:
	monthly = summarize_monthly(df, username, date_start=date_start, date_end=date_end)
	if monthly.empty:
		return pd.DataFrame()

	months = monthly["Month"].sort_values().tolist()
	rows = {"Username": username or os.getenv("GRAB_USERNAME", "unknown")}

	for idx, month in enumerate(months, start=1):
		month_data = monthly.loc[monthly["Month"].eq(month)].iloc[0]
		rows[f"Omzet Bulan ke-{idx}"] = float(month_data["Omzet_Net_Sales"])

	for idx, month in enumerate(months, start=1):
		month_data = monthly.loc[monthly["Month"].eq(month)].iloc[0]
		rows[f"Order Bulan ke-{idx}"] = int(month_data["Order_Count"])

	return pd.DataFrame([rows])


def print_summary(summary: pd.DataFrame) -> None:
	if summary.empty:
		print("Tidak ada data valid yang cocok dengan aturan filter.")
		return

	display = summary.copy()
	display["Month"] = display["Month"].dt.strftime("%Y-%m")
	display["Omzet_Net_Sales"] = display["Omzet_Net_Sales"].apply(format_rupiah)

	print(display.to_string(index=False))


def push_to_gsheet(username: str, wide_summary: pd.DataFrame, outlet: str = "", branch: str = "") -> None:
	url = "https://script.google.com/macros/s/AKfycbz8zCLNqDnVaz6Iau7uD-hZiynpaHigjtElk6Wlb5onr_Y9pRgfjtEkYm9unr1cNxkq/exec"
	if wide_summary.empty:
		return

	row = wide_summary.iloc[0]

	# Mapping berdasarkan kunci di wide_summary
	payload = {
		"username": str(username),
		"outlet": str(outlet),
		"branch": str(branch),
		"omzet1": float(row.get("Omzet Bulan ke-1", 0)),
		"omzet2": float(row.get("Omzet Bulan ke-2", 0)),
		"omzet3": float(row.get("Omzet Bulan ke-3", 0)),
		"order1": int(row.get("Order Bulan ke-1", 0)),
		"order2": int(row.get("Order Bulan ke-2", 0)),
		"order3": int(row.get("Order Bulan ke-3", 0)),
	}

	print(f"\nPushing data to Google Sheets for {username}...")
	try:
		# Apps Script requires follow redirects (default in requests)
		response = requests.post(url, json=payload, timeout=30)
		if response.status_code == 200:
			try:
				result = response.json()
				if result.get("status") == "success":
					print("✓ Berhasil dikirim ke Google Sheets!")
				else:
					print(f"✗ Gagal: {result.get('message')}")
			except:
				# Sometimes Apps Script returns HTML error page even with 200
				print("✓ Push terkirim (cek GSheet jika status tidak muncul)")
		else:
			print(f"✗ Gagal mengirim (HTTP {response.status_code})")
	except Exception as e:
		print(f"✗ Error saat push ke Google Sheets: {str(e)}")


def _build_output_path(
	args_output: str,
	date_start: "pd.Timestamp | None",
	date_end: "pd.Timestamp | None",
	username: str | None,
) -> Path:
	"""Tentukan path output.

	Jika user menyediakan --output eksplisit → pakai itu.
	Jika ada rentang tanggal → laporan/<start>_<end>/<username>.xlsx
	Fallback → DEFAULT_OUTPUT
	"""
	default_str = str(DEFAULT_OUTPUT)
	if args_output != default_str:
		# User override eksplisit
		return Path(args_output).expanduser().resolve()

	if date_start is not None:
		start_str = date_start.strftime("%Y-%m-%d")
		end_str = date_end.strftime("%Y-%m-%d") if date_end else "sekarang"
		folder = LAPORAN_DIR / f"{start_str}_{end_str}"
		safe_name = (username or "unknown").replace("/", "_").replace("\\", "_")
		return folder / f"{safe_name}.xlsx"

	return DEFAULT_OUTPUT


def main(username: str = None, outlet: str = "", branch: str = "") -> None:
	parser = argparse.ArgumentParser(
		description="Hitung omzet per bulan dan total order per bulan dari file Grab transactions.",
	)
	parser.add_argument("csv_path", nargs="?", help="Path file CSV transaksi")
	parser.add_argument(
		"--output",
		default=str(DEFAULT_OUTPUT),
		help="Path output XLSX/CSV (opsional, default otomatis ke folder laporan/)",
	)
	parser.add_argument(
		"--start-date",
		default=None,
		help="Filter awal (inklusif), format YYYY-MM-DD. Contoh: 2026-05-01",
	)
	parser.add_argument(
		"--end-date",
		default=None,
		help="Filter akhir (inklusif), format YYYY-MM-DD. Contoh: 2026-05-07",
	)
	args, _ = parser.parse_known_args()

	# Parse date filters
	try:
		date_start = pd.Timestamp(args.start_date) if args.start_date else None
		date_end = (
			pd.Timestamp(args.end_date).replace(hour=23, minute=59, second=59)
			if args.end_date
			else None
		)
	except Exception as exc:
		parser.error(f"Format tanggal tidak valid: {exc}")

	if date_start:
		print(f"Filter tanggal: {date_start.date()} s/d {date_end.date() if date_end else 'tidak dibatasi'}")

	input_path = resolve_input_path(args.csv_path)
	output_path = _build_output_path(args.output, date_start, date_end, username)

	df = load_transactions(input_path)
	summary = summarize_monthly(df, username, date_start=date_start, date_end=date_end)
	wide_summary = summarize_wide(df, username, date_start=date_start, date_end=date_end)
	
	if outlet:
		wide_summary.insert(0, "Outlet", outlet)
	if branch:
		wide_summary.insert(1, "Branch", branch)

	output_path.parent.mkdir(parents=True, exist_ok=True)
	if output_path.suffix.lower() == ".xlsx":
		with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
			wide_summary.to_excel(writer, index=False, sheet_name="Summary")

		from openpyxl import load_workbook
		from openpyxl.styles import Alignment, Font

		workbook = load_workbook(output_path)
		worksheet = workbook["Summary"]

		for cell in worksheet[1]:
			cell.font = Font(bold=True)
			cell.alignment = Alignment(horizontal="center", vertical="center")

		for row in worksheet.iter_rows(min_row=2):
			for cell in row:
				cell.alignment = Alignment(horizontal="center", vertical="center")

		for column in worksheet.columns:
			max_length = 0
			column_letter = column[0].column_letter
			for cell in column:
				cell_value = "" if cell.value is None else str(cell.value)
				max_length = max(max_length, len(cell_value))
			worksheet.column_dimensions[column_letter].width = max(max_length + 2, 18)

		workbook.save(output_path)
	else:
		wide_summary.to_csv(output_path, index=False)

	print_summary(summary)
	if not wide_summary.empty:
		print("\nFormat spreadsheet (Data Mentah):")
		with pd.option_context("display.max_columns", None, "display.width", 200):
			print(wide_summary.to_string(index=False))

	# Push ke Google Sheets
	# user_to_push = username or os.getenv("GRAB_USERNAME", "unknown")
	# push_to_gsheet(user_to_push, wide_summary, outlet, branch)

	print(f"\nRingkasan disimpan ke: {output_path}")


if __name__ == "__main__":
	main()
